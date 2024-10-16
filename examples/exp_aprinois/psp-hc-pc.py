"""Compute the principle components of the PSP-RS and healthy control dataset."""

from pathlib import Path

import jax.numpy as jnp
import nibabel
from medclassx.mask_trafo import mask_vector
from medclassx.pca import pca
from openpyxl import load_workbook

#########################PREPARE DATA ACCESS###########################################
# goal: obtain a list that contains all paths to the data

# path to the excel document
path = Path(r"data\aprinois_nuk_data\TAU-PM-PBB3_Scores_gb_18122023_jb.xlsx")

# opens the excel document
workbook = load_workbook(filename=path)

# selects one sheet, in this case aprinoia data
sheet = workbook.active

# assemble the paths to the data
root_dir = r"data\aprinois_nuk_data\Warped_TAU"

# collect the file name of the healthy controls
hcs_folder_names = [
    row[0].replace("-", "") # mismatch between excel name and filename
    for row in sheet.iter_rows(
        min_row=2, max_row=31, min_col=1, max_col=1, values_only=True
    )
]

# the full filenames of the healthy controls
hcs_paths = [
    Path(root_dir + r"\ana_swr" + n + "_PM-PBB3.nii") for n in hcs_folder_names
]

# collect the filename of the PSP group
psp_folder_names = [
    row[0].replace("-", "") # mismatch between excel name and filename
    for row in sheet.iter_rows(
        min_row=32, max_row=61, min_col=1, max_col=1, values_only=True
    )
]

# the full filenames of the psp patients
psp_paths = [
    Path(root_dir + r"\ana_swr" + n + "_PM-PBB3.nii") for n in psp_folder_names
]

# collect all data paths in a list
paths = hcs_paths + psp_paths

#############################PREPARE MASKING###########################################

# path to mask
mask_path = Path(r"data\aprinois_nuk_data\mask_for_scanvp.nii")

# load the mask with nibabel
nifti_mask = nibabel.load(mask_path)

# extract the numerical data
mask = nifti_mask.get_fdata()

# make it into a vector mask
mask = mask.reshape(-1)

#############################PREPARE DATASET###########################################

# load the images with nibabel
nifti_imgs = [nibabel.load(p) for p in paths]

# extract the numerical data as float64
img_data = [n.get_fdata() for n in nifti_imgs]

# convert to a single 4-d jax tensor of shape (batch, a, b, c)
img_data = jnp.array(img_data)

print(f"Data group shape: {jnp.shape(img_data)} and dtype {img_data.dtype}.")

# data shapes
n, a, b, c = jnp.shape(img_data)

# Reshape into matrix form
X = jnp.reshape(img_data, shape=(n, a * b * c))

# prepare unmasking
unmask = mask_vector(X[0], mask)[1]

# mask the data
X = jnp.array([mask_vector(x, mask)[0] for x in X])

# shift data to the range [1, upper_bound]
X = X - jnp.min(X) + 1

# log transform
X = jnp.log(X)

# column center
X -= jnp.mean(X, axis=0, keepdims=True)

# row center
X -= jnp.mean(X, axis=1, keepdims=True)

###############################COMPUTE PCA#############################################

# perform pca
transform, recover, singular_values, W = pca(X)

# Compute the latent data representation
T = transform(X)

print(f"Shape of the latent dataset: {T.shape}")

# save the transformed dataset
jnp.save(Path(r"examples\exp_aprinois\out\latent_data.npy"), T)

# get the principle components back to unmasked space
pcs = [unmask(w) for w in W.transpose()]

# reshape back to 3d space
pcs = [jnp.reshape(pc, shape=(a, b, c)) for pc in pcs]

# save as niftis for visual inspection
for i, pc in enumerate(pcs):
    # totally unclear to me if passing these headers makes any sense!
    pc_to_nifti = nibabel.Nifti1Image(
        pc,
        affine=nifti_imgs[1].affine,
        header=nifti_imgs[1].header,
    )

    nibabel.save(
        pc_to_nifti, 
        "examples\exp_aprinois\out\pc_" + f"{i+1}" + ".nii",
    )


# Is this the right z-score normalization? What is a sensible reference that 
# the principle components should be normalized to?
# repeat the same for z transformed principle components
pcs_zscore = [unmask((w - jnp.mean(w))/jnp.std(w)) for w in W.transpose()]

# reshape to 3d space
pcs_zscore = [jnp.reshape(pc_z, shape=(a, b, c)) for pc_z in pcs_zscore]

# save as niftis for visual inspection
for i, pc_z in enumerate(pcs_zscore):
    # totally unclear to me if passing these headers makes any sense!
    pc_to_nifti = nibabel.Nifti1Image(
        pc_z,
        affine=nifti_imgs[1].affine,
        header=nifti_imgs[1].header,
    )

    nibabel.save(pc_to_nifti, "examples\exp_aprinois\out\zscore_pc_" + f"{i+1}" + ".nii")