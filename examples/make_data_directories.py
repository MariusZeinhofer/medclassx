r"""Uses the data from the excel file to put the data nicely into folders.

We want 3 folders:
    - data\AD
    - data\HC
    - data\PSP
to contain the corresponding sets of nifti files.
"""

import shutil
from pathlib import Path

from openpyxl import load_workbook

print("Prepare data access.")

# path to the excel document
path = Path(r"data\aprinois_nuk_data\TAU-PM-PBB3_Scores_gb_18122023_jb.xlsx")

# opens the excel document
workbook = load_workbook(filename=path)

# selects one sheet, in this case aprinoia data
sheet = workbook.active

# assemble the paths to the data
root_dir = r"data\aprinois_nuk_data\Warped_TAU\\"

# in the 15th column, the rows 2-31 contain the filenames of the healthy controls
hcs_paths = [
    Path(root_dir + row[0].strip() + ".nii")
    for row in sheet.iter_rows(
        min_row=2, max_row=31, min_col=15, max_col=15, values_only=True
    )
]

# in the 15th column, the rows 32-61 contain the filenames of the psp patients
psp_paths = [
    Path(root_dir + row[0].strip() + ".nii")
    for row in sheet.iter_rows(
        min_row=32, max_row=61, min_col=15, max_col=15, values_only=True
    )
]

# in the 15th column, the rows 62-91 contain the filenames of the ad patients
ad_paths = [
    Path(root_dir + row[0].strip() + ".nii")
    for row in sheet.iter_rows(
        min_row=62, max_row=91, min_col=15, max_col=15, values_only=True
    )
]


# directory to save the data in
directory_hc = Path(r"data\HC")
directory_ad = Path(r"data\AD")
directory_psp = Path(r"data\PSP")

# make fresh directories
if directory_hc.exists():
    shutil.rmtree(directory_hc)
directory_hc.mkdir(parents=True, exist_ok=True)

if directory_ad.exists():
    shutil.rmtree(directory_ad)
directory_ad.mkdir(parents=True, exist_ok=True)

if directory_psp.exists():
    shutil.rmtree(directory_psp)
directory_psp.mkdir(parents=True, exist_ok=True)


# Copy files
for hc_path in hcs_paths:
    dest_path = directory_hc / hc_path.name
    shutil.copy(hc_path, dest_path)

for psp_path in psp_paths:
    dest_path = directory_psp / psp_path.name
    shutil.copy(psp_path, dest_path)

for ad_path in ad_paths:
    dest_path = directory_ad / ad_path.name
    shutil.copy(ad_path, dest_path)
