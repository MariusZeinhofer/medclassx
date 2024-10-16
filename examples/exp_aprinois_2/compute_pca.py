"""Performs PCA-regression on the PSP-RS and healthy control dataset.

This script transforms and normalizes the dataset and computes the principle
components. It saves the following files:

- the first `cut_off` principle components in image space. That means
    it unmasks the files and transforms them back to volumetric format and saves
    them in the nifti format. This is done both for the z-score normalized PCs and
    for the unnormalized ones. These files are for visual inspection mainly.

- For every principle component it saves a json file that contains the principle
    component in masked vector form. It does this both for z-score normalized values
    and non-normalized values. Other information contained in each json file (one for
    each principle component) is the "pc_number", the singular value "sv" the variance
    accounted for "vaf" (singular_value/sum of singular values), the "accuracy", the
    principle component "pc" and the z-score normalized principle component "z_pc".
    Keys for the json: {"pc_number", "sv", "vaf", "accuracy", "pc", "pcz"}.
"""

import json
from pathlib import Path

import jax.numpy as jnp
import nibabel
from medclassx.binary_pca_regression import binary_pca_regression
from medclassx.mask_trafo import mask_vector
from openpyxl import load_workbook

#########################PREPARE DATA ACCESS###########################################
# goal: obtain a list that contains all paths to the data

print("Prepare data access.")

# path to the excel document
path = Path(r"data\aprinois_nuk_data\TAU-PM-PBB3_Scores_gb_18122023_jb.xlsx")

# opens the excel document
workbook = load_workbook(filename=path)

# selects one sheet, in this case aprinoia data
sheet = workbook.active

# assemble the paths to the data
root_dir = r"data\aprinois_nuk_data\Warped_TAU"

# collect the file name of the healthy controls
hcs_folder_names = [
    row[0].replace("-", "") # mismatch between excel name and filename
    for row in sheet.iter_rows(
        min_row=2, max_row=31, min_col=1, max_col=1, values_only=True
    )
]

# the full filenames of the healthy controls
hcs_paths = [
    Path(root_dir + r"\ana_swr" + n + "_PM-PBB3.nii") for n in hcs_folder_names
]

# collect the filename of the PSP group
psp_folder_names = [
    row[0].replace("-", "") # mismatch between excel name and filename
    for row in sheet.iter_rows(
        min_row=32, max_row=61, min_col=1, max_col=1, values_only=True
    )
]

# the full filenames of the psp patients
psp_paths = [
    Path(root_dir + r"\ana_swr" + n + "_PM-PBB3.nii") for n in psp_folder_names
]

# collect all data paths in a list
paths = hcs_paths + psp_paths

#############################PREPARE MASKING###########################################

print("Load mask.")

# path to mask
mask_path = Path(r"data\aprinois_nuk_data\mask_for_scanvp.nii")

# load the mask with nibabel
nifti_mask = nibabel.load(mask_path)

# extract the numerical data
mask = nifti_mask.get_fdata()

# make it into a vector mask
mask = mask.reshape(-1)

#############################PREPARE DATASET###########################################

print("Mask, shift, log tranform and double center data.")

# load the images with nibabel
nifti_imgs = [nibabel.load(p) for p in paths]

# extract the numerical data as float64
img_data = [n.get_fdata() for n in nifti_imgs]

# convert to a single 4-d jax tensor of shape (batch, a, b, c)
img_data = jnp.array(img_data)

# data shapes
n, a, b, c = jnp.shape(img_data)

# Reshape into matrix form
X = jnp.reshape(img_data, shape=(n, a * b * c))

# prepare unmasking
unmask = mask_vector(X[0], mask)[1]

# mask the data
X = jnp.array([mask_vector(x, mask)[0] for x in X])

# shift data to the range [1, upper_bound]
X = X - jnp.min(X) + 1

# log transform
X = jnp.log(X)

# column center
X -= jnp.mean(X, axis=0, keepdims=True)

# row center
X -= jnp.mean(X, axis=1, keepdims=True)

# save for later use
jnp.save(Path(r"examples\exp_aprinois_2\out\X.npy"), X)

print(f"Done, data of shape: {X.shape}.")

#################################PCA REGRESSION########################################

print("Compute PCA regression.")

# the first 30 are healthy controls
X_con = X[0:30]

# the remaining 30 are PSP patients
X_pat = X[30:]

# compute the pca regression
results, T = binary_pca_regression(X_control=X_con, X_patient=X_pat, cut_off=20)

print("Done, save results.")

##################################SAVE RESULTS#########################################

# save the transformed dataset
jnp.save(Path(r"examples\exp_aprinois_2\out\latent_data.npy"), T)

# retrieve the principle components from the results list
pcs = [result["pc"] for result in results]

# get the principle components back to unmasked space
pcs = [unmask(pc) for pc in pcs]

# reshape back to 3d space
pcs = [jnp.reshape(pc, shape=(a, b, c)) for pc in pcs]

# save as niftis for visual inspection
for i, pc in enumerate(pcs):
    # totally unclear to me if passing these headers makes any sense!
    pc_to_nifti = nibabel.Nifti1Image(
        pc,
        affine=nifti_imgs[1].affine,
        header=nifti_imgs[1].header,
    )

    nibabel.save(
        pc_to_nifti, 
        Path("examples\exp_aprinois_2\out\pc_" + f"{i+1}" + ".nii"),
    )

# repeat to save the z-score normalized principle components
# retrieve the normalized principle components from the results list
pcs_zscore = [d["pcz"] for d in results]

# get the principle components back to unmasked space
pcs_zscore = [unmask(pcz) for pcz in pcs_zscore]

# reshape back to 3d space
pcs_zscore = [jnp.reshape(pcz, shape=(a, b, c)) for pcz in pcs_zscore]

# save as niftis for visual inspection
for i, pcz in enumerate(pcs_zscore):
    # totally unclear to me if passing these headers makes any sense!
    pcz_to_nifti = nibabel.Nifti1Image(
        pcz,
        affine=nifti_imgs[1].affine,
        header=nifti_imgs[1].header,
    )

    nibabel.save(
        pcz_to_nifti, 
        Path("examples\exp_aprinois_2\out\zscore_pc_" + f"{i+1}" + ".nii"),
    )

# save principle components and metadata as json file
for i, d in enumerate(results):
    
    # jax and numpy arrays cannot be saved in a json file, use list instead
    d_ser = {k: (v.tolist() if hasattr(v, "tolist") else v) for k, v in d.items()}

    with open(Path(r"examples\exp_aprinois_2\out\pc_" + f"{i+1}.json"), "w") as j_file:
        json.dump(d_ser, j_file)